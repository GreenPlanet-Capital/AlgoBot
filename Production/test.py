# importing openpyxl module
import openpyxl
  
# Give the location of the file
path = "C:\\Users\\Sarbajit\\Code\\AlgoBot\\S&P500_list.xlsx"
  
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
  
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
  
# Loop will print all values
# of first column 
outlist = []
for i in range(4, 503):
    cell_obj = sheet_obj.cell(row = i, column = 3)
    outlist.append(cell_obj.value)
print(outlist)